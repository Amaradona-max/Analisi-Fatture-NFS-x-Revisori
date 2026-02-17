from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.services.file_processor import CompareFTFileProcessor, NFSFTFileProcessor


@pytest.fixture
def sample_dataframe():
    return pd.DataFrame(
        {
            "C_NOME": ["ACME Inc", "Test Corp", "ACME Inc"],
            "FAT_DATDOC": ["2025-01-01", "2025-01-02", "2025-01-01"],
            "FAT_NDOC": ["F001", "F002", "F001"],
            "FAT_DATREG": ["2025-01-01", "2025-01-02", "2025-01-01"],
            "FAT_PROT": ["EP", "P", "EP"],
            "FAT_NUM": [1, 2, 1],
            "IMPONIBILE": [100.0, 200.0, 100.0],
            "FAT_TOTFAT": [122.0, 244.0, 122.0],
            "FAT_TOTIVA": [22.0, 44.0, 22.0],
            "DMA_NUM": ["M001", "M002", "M001"],
            "RA_CODTRIB": ["I9", "XX", "RO"],
            "RA_IMPOSTA": [5.0, 10.0, 5.0],
            "RA_IMPON": [50.0, 100.0, 50.0],
            "TMA_TOT": [122.0, 244.0, 122.0],
            "TMC_G8": ["ID1", "ID2", "ID1"],
        }
    )


def test_validate_file_success(sample_dataframe):
    processor = NFSFTFileProcessor()
    processor.validate_file(sample_dataframe)


def test_validate_file_missing_columns():
    processor = NFSFTFileProcessor()
    df = pd.DataFrame({"WRONG_COL": [1, 2]})
    with pytest.raises(ValueError, match="Colonne mancanti"):
        processor.validate_file(df)


def test_process_file_removes_duplicates(sample_dataframe, tmp_path: Path):
    processor = NFSFTFileProcessor()
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    sample_dataframe.to_excel(input_path, index=False)
    stats = processor.process_file(input_path, output_path)

    assert output_path.exists()
    assert stats["total_records"] == 2
    assert stats["duplicates_removed"] == 1
    assert stats["fase2_records"] == 1
    assert stats["fase3_records"] == 1


def test_process_compare_creates_output(tmp_path: Path):
    nfs_path = tmp_path / "nfs.xlsx"
    pisa_path = tmp_path / "pisa.xlsx"
    out_path = tmp_path / "out.xlsx"

    pd.DataFrame(
        {
            "C_NOME": ["A", "B", "C"],
            "FAT_PROT": ["P", "EP", "EP"],
            "FAT_NUM": [1, 2, 3],
            "FAT_NDOC": ["1", "2", "3"],
            "FAT_DATDOC": ["2025-01-05", "2025-01-10", "2025-02-05"],
            "FAT_DATREG": ["2025-01-06", "2025-01-11", "2025-02-05"],
            "FAT_TOTIVA": [10, 20, 5],
            "IMPONIBILE": [100, 200, 10],
            "FAT_TOTFAT": [110, 220, 12],
            "RA_IMPON": [0, 0, 0],
            "RA_IMPOSTA": [0, 0, 0],
            "RA_CODTRIB": ["", "", ""],
            "TMC_G8": ["", "123", "999"],
        }
    ).to_excel(nfs_path, index=False)

    pd.DataFrame(
        {
            "Identificativo SDI": ["", "123", "999"],
            "Creditore": ["X", "Y", "Z"],
            "Numero fattura": ["PX-1", "PY-1", "PZ-1"],
            "Data emissione": ["2025-01-07", "2025-01-08", "2025-01-09"],
            "Importo fattura": ["50", "70", "10"],
        }
    ).to_excel(pisa_path, index=False)

    processor = CompareFTFileProcessor()
    summary = processor.process_files(nfs_path, pisa_path, out_path)

    assert out_path.exists()
    assert summary["period"] == "2025-01"
    assert summary["nfs"]["cartacee"]["count"] == 1
    assert summary["nfs"]["cartacee"]["amount"] == 100.0
    assert summary["nfs"]["elettroniche"]["count"] == 1
    assert summary["nfs"]["elettroniche"]["amount"] == 200.0
    assert summary["pisa"]["cartacee"]["count"] == 1
    assert summary["pisa"]["cartacee"]["amount"] == 50.0

    wb = load_workbook(out_path)
    assert "Confronto" in wb.sheetnames
    assert "Fatture da Verificare" in wb.sheetnames
    assert "Differenze Elettroniche SDI" in wb.sheetnames
    assert "Differenze SDI in Comune" in wb.sheetnames
    assert "Pisa Solo - Mese NFS" in wb.sheetnames
    ws = wb["Fatture da Verificare"]
    assert ws.max_row >= 3

    ws_diff = wb["Differenze SDI in Comune"]
    assert ws_diff.max_row >= 2
    assert ws_diff["A2"].value == "123"

    ws_pisa = wb["Pisa Solo - Mese NFS"]
    found = False
    for r in range(2, ws_pisa.max_row + 1):
        if ws_pisa.cell(r, 1).value == "999":
            months = ws_pisa.cell(r, 6).value or ""
            assert "2025-02" in months
            found = True
            break
    assert found
