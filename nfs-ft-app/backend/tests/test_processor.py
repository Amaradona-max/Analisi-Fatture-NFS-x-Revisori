from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
import pytest

from app.services.file_processor import NFSFTFileProcessor, PisaFTFileProcessor, CompareFTFileProcessor


@pytest.fixture
def sample_dataframe():
    return pd.DataFrame(
        {
            "C_NOME": ["ACME Inc", "Test Corp", "ACME Inc"],
            "FAT_DATDOC": ["2025-01-01", "2025-01-02", "2025-01-01"],
            "FAT_NDOC": ["F001", "F002", "F001"],
            "FAT_DATREG": ["2025-01-01", "2025-01-02", "2025-01-01"],
            "FAT_PROT": ["EP", "P", "EP"],
            "FAT_NUM": [1, 2, 1],
            "IMPONIBILE": [100.0, 200.0, 100.0],
            "FAT_TOTFAT": [122.0, 244.0, 122.0],
            "FAT_TOTIVA": [22.0, 44.0, 22.0],
            "RA_IMPON": [100.0, 200.0, 100.0],
            "RA_CODTRIB": ["I9", "XX", "RO"],
            "RA_IMPOSTA": [5.0, 10.0, 5.0],
            "TMC_G8": ["ID1", "ID2", "ID1"],
        }
    )


def test_validate_file_success(sample_dataframe):
    processor = NFSFTFileProcessor()
    processor.validate_file(sample_dataframe)


def test_validate_file_missing_columns():
    processor = NFSFTFileProcessor()
    df = pd.DataFrame({"WRONG_COL": [1, 2]})
    with pytest.raises(ValueError, match="Colonne mancanti"):
        processor.validate_file(df)


def test_process_file_removes_duplicates(sample_dataframe, tmp_path: Path):
    processor = NFSFTFileProcessor()
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    sample_dataframe.to_excel(input_path, index=False)
    stats = processor.process_file(input_path, output_path)

    assert output_path.exists()
    assert stats["total_records"] == 2
    assert stats["duplicates_removed"] == 1
    assert stats["fase2_records"] == 1
    assert stats["fase3_records"] == 1


def test_process_file_pisa_splits_by_sdi(tmp_path: Path):
    columns = [
        "Identificativo SDI",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "Creditore",
        "I",
        "Importo Fattura",
        "K",
        "Importo Pagato",
        "M",
        "N",
        "O",
    ]
    df = pd.DataFrame(
        [
            ["", "b1", "c1", "d1", "e1", "2025-01-10", "g1", "Ragione A", "i1", 120.0, "k1", 100.0, "m1", "n1", "o1"],
            ["123", "b2", "c2", "d2", "e2", "", "g2", "Ragione B", "i2", 220.0, "k2", 200.0, "m2", "n2", "o2"],
            [None, "b3", "c3", "d3", "e3", "2025-02-01", "g3", "Ragione C", "i3", 320.0, "k3", 300.0, "m3", "n3", "o3"],
        ],
        columns=columns,
    )

    processor = PisaFTFileProcessor()
    input_path = tmp_path / "input_pisa.xlsx"
    output_path = tmp_path / "output_pisa.xlsx"
    df.to_excel(input_path, index=False)

    stats = processor.process_file(input_path, output_path)

    assert output_path.exists()
    assert stats["total_records"] == 2
    assert stats["fase2_records"] == 2
    assert stats["fase3_records"] == 0

    wb = load_workbook(output_path, data_only=True)
    cartacee_ws = wb["Fatture Cartacee"]
    elettroniche_ws = wb["Fatture Elettroniche"]

    assert cartacee_ws["A1"].value == "NUMERO TOTALE"
    assert cartacee_ws["B1"].value == "IMPONIBILE"
    assert cartacee_ws["A2"].value == 2
    assert cartacee_ws["B2"].value == 400.0

    assert elettroniche_ws["A1"].value == "NUMERO TOTALE"
    assert elettroniche_ws["B1"].value == "IMPONIBILE"
    assert elettroniche_ws["A2"].value == 0


def test_compare_files_january_2025(tmp_path: Path):
    nfs_df = pd.DataFrame(
        {
            "C_NOME": ["A", "B"],
            "FAT_DATDOC": ["2025-01-05", "2025-02-05"],
            "FAT_NDOC": ["F001", "F002"],
            "FAT_DATREG": ["2025-01-10", "2025-02-10"],
            "FAT_PROT": ["P", "EP"],
            "FAT_NUM": [1, 2],
            "IMPONIBILE": [100.0, 200.0],
            "FAT_TOTFAT": [122.0, 244.0],
            "FAT_TOTIVA": [22.0, 44.0],
            "RA_IMPON": [100.0, 200.0],
            "RA_CODTRIB": ["I9", "RO"],
            "RA_IMPOSTA": [5.0, 10.0],
            "TMC_G8": ["", "ID2"],
        }
    )

    pisa_columns = [
        "Identificativo SDI",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "Creditore",
        "I",
        "Importo Fattura",
        "K",
        "Importo Pagato",
        "M",
        "N",
        "O",
    ]
    pisa_df = pd.DataFrame(
        [
            ["", "b1", "c1", "d1", "e1", "2025-01-12", "g1", "Ragione A", "i1", 120.0, "k1", 150.0, "m1", "n1", "o1"],
            ["123", "b2", "c2", "d2", "e2", "2025-01-20", "g2", "Ragione B", "i2", 220.0, "k2", 250.0, "m2", "n2", "o2"],
            ["", "b3", "c3", "d3", "e3", "2025-02-05", "g3", "Ragione C", "i3", 320.0, "k3", 300.0, "m3", "n3", "o3"],
        ],
        columns=pisa_columns,
    )

    nfs_path = tmp_path / "nfs.xlsx"
    pisa_path = tmp_path / "pisa.xlsx"
    output_path = tmp_path / "compare.xlsx"
    nfs_df.to_excel(nfs_path, index=False)
    pisa_df.to_excel(pisa_path, index=False)

    processor = CompareFTFileProcessor()
    summary = processor.process_files(nfs_path, pisa_path, output_path)

    assert output_path.exists()
    assert summary["nfs"]["cartacee"]["count"] == 1
    assert summary["nfs"]["elettroniche"]["count"] == 1
    assert summary["pisa"]["cartacee"]["count"] == 1
    assert summary["pisa"]["elettroniche"]["count"] == 1

    wb = load_workbook(output_path, data_only=True)
    assert "Diffrenze da Verificare" in wb.sheetnames
    diff_ws = wb["Diffrenze da Verificare"]
    assert diff_ws.max_row == 2
