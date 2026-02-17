from pathlib import Path
from typing import Any, Dict, Optional
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


logger = logging.getLogger(__name__)


class NFSFTFileProcessor:
    PROTOCOLLI_FASE2 = ["P", "2P", "LABI", "FCBI", "FCSI", "FCBE", "FCSE"]
    PROTOCOLLI_FASE3 = [
        "EP",
        "2EP",
        "EL",
        "2EL",
        "EZ",
        "2EZ",
        "EZP",
        "FPIC",
        "FSIC",
        "FPEC",
        "FSEC",
        "AFIC",
        "ASIC",
        "AFEC",
        "ASEC",
        "ACBI",
        "ACSI",
        "ACBE",
        "ACSE",
    ]

    DESCRIZIONI_FASE2 = {
        "P": "Fatture Cartacee San",
        "2P": "Fatture Cartacee Ter",
        "LABI": "Fatture Lib.Prof. San",
        "FCBI": "Fatture Cartacee Estere San",
        "FCSI": "Fatture Cartacee Estere San",
        "FCBE": "Fatture Cartacee Estere San",
        "FCSE": "Fatture Cartacee Estere San",
    }

    DESCRIZIONI_FASE3 = {
        "EP": "Fatture Elettroniche San",
        "2EP": "Fatture Elettroniche Ter",
        "EL": "Fatture Elettroniche Lib.Prof. San",
        "2EL": "Fatture Elettroniche Lib.Prof. Ter",
        "EZ": "Fatture Elettroniche Commerciali San",
        "2EZ": "Fatture Elettroniche Commerciali Ter",
        "EZP": "Fatture Elettroniche Commerciali San",
        "FPIC": "Fatture Elettroniche Estere San",
        "FSIC": "Fatture Elettroniche Estere San",
        "FPEC": "Fatture Elettroniche Estere San",
        "FSEC": "Fatture Elettroniche Estere San",
        "AFIC": "Fatture Elettroniche Estere San",
        "ASIC": "Fatture Elettroniche Estere San",
        "AFEC": "Fatture Elettroniche Estere San",
        "ASEC": "Fatture Elettroniche Estere San",
        "ACBI": "Fatture Elettroniche Estere San",
        "ACSI": "Fatture Elettroniche Estere San",
        "ACBE": "Fatture Elettroniche Estere San",
        "ACSE": "Fatture Elettroniche Estere San",
    }

    REQUIRED_COLUMNS = [
        "C_NOME",
        "FAT_DATDOC",
        "FAT_NDOC",
        "FAT_DATREG",
        "FAT_PROT",
        "FAT_NUM",
        "IMPONIBILE",
        "FAT_TOTFAT",
        "FAT_TOTIVA",
        "RA_IMPON",
        "RA_CODTRIB",
        "RA_IMPOSTA",
        "TMC_G8",
    ]

    def __init__(self) -> None:
        self.all_protocols = self.PROTOCOLLI_FASE2 + self.PROTOCOLLI_FASE3

    def validate_file(self, df: pd.DataFrame) -> None:
        missing_cols = [col for col in self.REQUIRED_COLUMNS if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Colonne mancanti: {', '.join(missing_cols)}")

    def _filter_january_2025(self, df: pd.DataFrame, date_column: str) -> pd.DataFrame:
        if date_column not in df.columns:
            return df.iloc[0:0].copy()
        date_series = pd.to_datetime(df[date_column], errors="coerce")
        start = pd.Timestamp(year=2025, month=1, day=1)
        end = pd.Timestamp(year=2025, month=1, day=31)
        mask = date_series.between(start, end)
        return df[mask].copy()

    def process_file(self, input_path: Path, output_path: Path) -> Dict[str, Any]:
        try:
            logger.info("Caricamento file: %s", input_path)
            df = pd.read_excel(input_path)

            self.validate_file(df)

            df["FAT_PROT"] = df["FAT_PROT"].astype(str).str.strip()
            totale_iniziale = len(df)
            df_senza_duplicati = df.drop_duplicates(subset=["FAT_NUM", "C_NOME"]).copy()
            duplicati_rimossi = totale_iniziale - len(df_senza_duplicati)
            df_filtrato = df_senza_duplicati[df_senza_duplicati["FAT_PROT"].isin(self.all_protocols)].copy()

            if len(df_filtrato) == 0:
                raise ValueError("Nessun protocollo valido trovato nel file")

            df_filtrato["RA_CODTRIB"] = (
                df_filtrato["RA_CODTRIB"]
                .astype(str)
                .str.strip()
                .where(lambda value: value.isin(["I9", "RO"]), "")
            )

            colonne_ordinate = [
                "C_NOME",
                "FAT_DATDOC",
                "FAT_NDOC",
                "FAT_DATREG",
                "FAT_PROT",
                "FAT_NUM",
                "FAT_TOTIVA",
                "IMPONIBILE",
                "FAT_TOTFAT",
                "RA_CODTRIB",
                "RA_IMPOSTA",
                "RA_IMPON",
                "TMC_G8",
            ]

            df_finale = df_filtrato[colonne_ordinate].copy()
            df_finale.columns = [
                "Ragione Sociale",
                "Data Fatture",
                "N. Fatture",
                "Data Registrazione",
                "Protocollo",
                "N. Protocollo",
                "Imposta",
                "Tot. Imponibile",
                "Tot. Imp. Fatture",
                "Rit. Codice Tributo",
                "Rit. Imposta",
                "Rit. Imp.",
                "Identificativo SDI",
            ]

            df_finale["Data Fatture"] = pd.to_datetime(df_finale["Data Fatture"], errors="coerce")
            df_finale["Data Registrazione"] = pd.to_datetime(df_finale["Data Registrazione"], errors="coerce")

            df_finale = df_finale.sort_values("Data Registrazione")

            df_dati = df_finale.copy()
            if "Data Registrazione" in df_dati.columns:
                df_dati["Imponibile"] = df_dati["Data Registrazione"]
                df_dati = df_dati.drop(columns=["Data Registrazione"], errors="ignore")
            df_dati = df_dati.drop(columns=["Tot. Imponibile"], errors="ignore")
            ordered_columns = [
                "Ragione Sociale",
                "Data Fatture",
                "N. Fatture",
                "Protocollo",
                "N. Protocollo",
                "Imposta",
                "Imponibile",
                "Tot. Imp. Fatture",
                "Rit. Codice Tributo",
                "Rit. Imposta",
                "Rit. Imp.",
                "Identificativo SDI",
            ]
            df_dati = df_dati[[col for col in ordered_columns if col in df_dati.columns]]

            stats = self._calculate_stats(df_finale, duplicati_rimossi)
            self._create_excel_output(df_finale, output_path, display_df=df_dati)

            logger.info("File elaborato con successo: %s", stats)
            return stats
        except Exception as exc:
            logger.error("Errore elaborazione file: %s", str(exc))
            raise

    def _calculate_stats(self, df: pd.DataFrame, duplicates_removed: int) -> Dict[str, Any]:
        fase2_count = df[df["Protocollo"].isin(self.PROTOCOLLI_FASE2)].shape[0]
        fase3_count = df[df["Protocollo"].isin(self.PROTOCOLLI_FASE3)].shape[0]

        return {
            "total_records": len(df),
            "fase2_records": fase2_count,
            "fase3_records": fase3_count,
            "duplicates_removed": duplicates_removed,
            "protocols_fase2": self._count_by_protocol(df, self.PROTOCOLLI_FASE2),
            "protocols_fase3": self._count_by_protocol(df, self.PROTOCOLLI_FASE3),
        }

    def _count_by_protocol(self, df: pd.DataFrame, protocols: list) -> Dict[str, int]:
        counts = {}
        for prot in protocols:
            counts[prot] = len(df[df["Protocollo"] == prot])
        return counts

    def _create_excel_output(
        self,
        df: pd.DataFrame,
        output_path: Path,
        display_df: Optional[pd.DataFrame] = None,
    ) -> None:
        wb = Workbook()

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True)

        ws_dati = self._add_dataframe_sheet(
            wb,
            "Dati",
            display_df if display_df is not None else df,
            header_fill,
            header_font,
            total_fill,
            total_font,
            date_columns=["Data Fatture", "Data Registrazione", "Imponibile"],
            money_columns=["Imposta", "Tot. Imponibile", "Tot. Imp. Fatture", "Rit. Imposta", "Rit. Imp."],
            use_active=True,
        )

        ws_nota2 = wb.create_sheet("Fatture Cartacee")
        self._create_summary_sheet(
            ws_nota2,
            df,
            self.PROTOCOLLI_FASE2,
            self.DESCRIZIONI_FASE2,
            header_fill,
            header_font,
                total_fill,
                total_font,
        )

        ws_nota3 = wb.create_sheet("Fatture Elettroniche")
        self._create_summary_sheet(
            ws_nota3,
            df,
            self.PROTOCOLLI_FASE3,
            self.DESCRIZIONI_FASE3,
            header_fill,
            header_font,
                total_fill,
                total_font,
        )

        wb.save(output_path)

    def _add_dataframe_sheet(
        self,
        wb: Workbook,
        title: str,
        df: pd.DataFrame,
        header_fill: PatternFill,
        header_font: Font,
        total_fill: PatternFill,
        total_font: Font,
        date_columns=None,
        money_columns=None,
        date_format: str = "mm/dd/yyyy",
        add_totals: bool = True,
        auto_size: bool = True,
        use_active: bool = False,
    ):
        ws = wb.active if use_active else wb.create_sheet(title)
        ws.title = title

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if auto_size:
            sample_rows = 50
            max_row = min(ws.max_row, sample_rows + 1)
            for column in ws.iter_cols(max_row=max_row):
                max_len = max((len(str(c.value or "")) for c in column), default=8)
                letter = column[0].column_letter
                ws.column_dimensions[letter].width = min(max_len + 2, 45)

        date_columns = date_columns or []
        money_columns = money_columns or []
        money_columns = [column for column in money_columns if column in df.columns]
        header_index = {cell.value: cell.column for cell in ws[1]}
        money_format = "#,##0.00"

        for column_name in date_columns:
            column_index = header_index.get(column_name)
            if column_index:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=column_index, max_col=column_index):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = date_format

        for column_name in money_columns:
            column_index = header_index.get(column_name)
            if column_index:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=column_index, max_col=column_index):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = money_format

        if money_columns and add_totals:
            totals = {}
            for column_name in money_columns:
                totals[column_name] = pd.to_numeric(df[column_name], errors="coerce").sum()
            total_row = ["TOTALE"] + [""] * (len(df.columns) - 1)
            for column_name, total_value in totals.items():
                total_row[df.columns.get_loc(column_name)] = total_value
            ws.append(total_row)
            total_row_index = ws.max_row
            for cell in ws[total_row_index]:
                cell.fill = total_fill
                cell.font = total_font
            for column_name in money_columns:
                column_index = header_index.get(column_name)
                if column_index:
                    ws.cell(row=total_row_index, column=column_index).number_format = money_format

        return ws

    def _create_summary_sheet(self, ws, df, protocols, descriptions, header_fill, header_font, total_fill, total_font):
        ws["A1"] = "PROTOCOLLO"
        ws["B1"] = "DESCRIZIONE"
        ws["C1"] = "NUMERO TOTALE"
        ws["D1"] = "IMPONIBILE"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        money_format = "#,##0.00"
        row = 2
        for prot in protocols:
            count = len(df[df["Protocollo"] == prot])
            imponibile_totale = pd.to_numeric(
                df.loc[df["Protocollo"] == prot, "Tot. Imponibile"], errors="coerce"
            ).sum()
            ws[f"A{row}"] = prot
            ws[f"B{row}"] = descriptions[prot]
            ws[f"C{row}"] = count
            ws[f"D{row}"] = imponibile_totale
            ws[f"D{row}"].number_format = money_format
            row += 1

        ws[f"A{row}"] = "TOTALE"
        ws[f"A{row}"].font = total_font
        ws[f"C{row}"] = f"=SUM(C2:C{row - 1})"
        ws[f"C{row}"].font = total_font
        ws[f"D{row}"] = f"=SUM(D2:D{row - 1})"
        ws[f"D{row}"].number_format = money_format
        for cell in ws[row]:
            cell.fill = total_fill
            cell.font = total_font

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20


class PisaFTFileProcessor(NFSFTFileProcessor):
    SELECTED_LETTERS = ["H", "C", "D", "E", "F", "O", "L", "J", "A"]
    RENAME_MAP = {
        "H": "Ragione Sociale",
        "L": "Imponibile",
        "J": "Imp.Tot. Fatture",
    }
    MONEY_COLUMNS = ["Imponibile", "Imp.Tot. Fatture"]
    USECOLS_RANGE = "A:O"
    MAX_DETAIL_ROWS = 5000

    def process_file(self, input_path: Path, output_path: Path) -> Dict[str, Any]:
        try:
            logger.info("Caricamento file Pisa Pagato: %s", input_path)
            df = pd.read_excel(input_path, usecols=self.USECOLS_RANGE, dtype=str)

            required_indices = self._letters_to_indices(self.SELECTED_LETTERS)
            max_index = max(required_indices)
            if df.shape[1] <= max_index:
                missing_letters = [
                    letter for letter, index in zip(self.SELECTED_LETTERS, required_indices) if index >= df.shape[1]
                ]
                raise ValueError(f"Colonne mancanti: {', '.join(missing_letters)}")

            data_pagamento_column = df.columns[self._letters_to_indices(["F"])[0]]
            pagamento_series = df[data_pagamento_column]
            pagamento_mask = ~(pagamento_series.isna() | (pagamento_series.astype(str).str.strip() == ""))
            df_pagato = df[pagamento_mask].copy()

            selected_indices = self._letters_to_indices(self.SELECTED_LETTERS)
            selected_columns = []
            for letter, index in zip(self.SELECTED_LETTERS, selected_indices):
                selected_columns.append(self.RENAME_MAP.get(letter) or df_pagato.columns[index])

            df_finale = df_pagato.iloc[:, selected_indices].copy()
            df_finale.columns = selected_columns
            data_pagamento_column_name = selected_columns[self.SELECTED_LETTERS.index("F")]
            df_finale = self._filter_january_2025(df_finale, data_pagamento_column_name)

            sdi_column = df.columns[self._letters_to_indices(["A"])[0]]
            cartacee_df, elettroniche_df = self._split_by_sdi(df_finale, sdi_column)
            df_dati = self._build_pisa_dati(df_finale)
            self._create_excel_output(df_finale, cartacee_df, elettroniche_df, output_path, display_df=df_dati)
            stats = {
                "total_records": len(df_finale),
                "fase2_records": len(cartacee_df),
                "fase3_records": len(elettroniche_df),
                "duplicates_removed": 0,
                "protocols_fase2": {"Cartacee": len(cartacee_df)},
                "protocols_fase3": {"Elettroniche": len(elettroniche_df)},
            }
            logger.info("File Pisa Pagato elaborato con successo: %s", stats)
            return stats
        except Exception as exc:
            logger.error("Errore elaborazione file Pisa Pagato: %s", str(exc))
            raise

    def _create_excel_output(
        self,
        df: pd.DataFrame,
        cartacee_df: pd.DataFrame,
        elettroniche_df: pd.DataFrame,
        output_path: Path,
        display_df: Optional[pd.DataFrame] = None,
    ) -> None:
        wb = Workbook()

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True)

        dati_df = display_df if display_df is not None else df
        self._add_dataframe_sheet(
            wb,
            "Dati",
            dati_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
            date_columns=[column for column in dati_df.columns if "data" in str(column).lower()],
            date_format="dd/mm/yyyy",
            money_columns=[
                column
                for column in ("Imposta", "Imponibile", "Tot. Imp. Fatture", "Rit. Imposta", "Rit. Imp.")
                if column in dati_df.columns
            ],
            auto_size=False,
            use_active=True,
        )

        ws_cartacee = wb.create_sheet("Fatture Cartacee")
        self._create_simple_summary_sheet(
            ws_cartacee,
            cartacee_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        ws_elettroniche = wb.create_sheet("Fatture Elettroniche")
        self._create_simple_summary_sheet(
            ws_elettroniche,
            elettroniche_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        wb.save(output_path)

    def _build_pisa_dati(self, df: pd.DataFrame) -> pd.DataFrame:
        selected_columns = list(df.columns)
        col_creditore = selected_columns[0] if len(selected_columns) > 0 else None
        col_c = selected_columns[1] if len(selected_columns) > 1 else None
        col_d = selected_columns[2] if len(selected_columns) > 2 else None
        col_e = selected_columns[3] if len(selected_columns) > 3 else None
        col_f = selected_columns[4] if len(selected_columns) > 4 else None
        col_o = selected_columns[5] if len(selected_columns) > 5 else None
        col_imponibile = selected_columns[6] if len(selected_columns) > 6 else None
        col_tot_fatture = selected_columns[7] if len(selected_columns) > 7 else None
        col_sdi = selected_columns[8] if len(selected_columns) > 8 else None

        df_dati = pd.DataFrame(
            {
                "Ragione Sociale": df[col_creditore] if col_creditore else "",
                "Data Fatture": df[col_c] if col_c else "",
                "N. Fatture": df[col_d] if col_d else "",
                "Protocollo": df[col_e] if col_e else "",
                "N. Protocollo": df[col_f] if col_f else "",
                "Imposta": df[col_o] if col_o else "",
                "Imponibile": df[col_imponibile] if col_imponibile else "",
                "Tot. Imp. Fatture": df[col_tot_fatture] if col_tot_fatture else "",
                "Rit. Codice Tributo": "",
                "Rit. Imposta": "",
                "Rit. Imp.": "",
                "Identificativo SDI": df[col_sdi] if col_sdi else "",
            }
        )
        return df_dati

    def _split_by_sdi(self, df: pd.DataFrame, sdi_column: str) -> tuple[pd.DataFrame, pd.DataFrame]:
        sdi_series = df[sdi_column]
        empty_mask = sdi_series.isna() | (sdi_series.astype(str).str.strip() == "")
        cartacee_df = df[empty_mask].copy()
        elettroniche_df = df[~empty_mask].copy()
        return cartacee_df, elettroniche_df

    def _letters_to_indices(self, letters: list[str]) -> list[int]:
        return [ord(letter) - ord("A") for letter in letters]

    def _create_simple_summary_sheet(
        self,
        ws,
        df: pd.DataFrame,
        header_fill: PatternFill,
        header_font: Font,
        total_fill: PatternFill,
        total_font: Font,
    ) -> None:
        ws["A1"] = "NUMERO TOTALE"
        ws["B1"] = "IMPONIBILE"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        imponibile_totale = pd.to_numeric(df["Imponibile"], errors="coerce").sum()
        ws["A2"] = len(df)
        ws["B2"] = imponibile_totale
        ws["B2"].number_format = "#,##0.00"

        for cell in ws[2]:
            cell.fill = total_fill
            cell.font = total_font

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20


class PisaRicevuteFTFileProcessor(NFSFTFileProcessor):
    PHASE = 1
    INPUT_REQUIRED_COLUMNS = [
        "Creditore",
        "Numero fattura",
        "Data emissione",
        "Data documento",
        "Data pagamento",
        "IVA",
        "Importo fattura",
        "Identificativo SDI",
    ]
    OUTPUT_COLUMNS = [
        "Ragione sociale",
        "N.fatture",
        "Data emissione",
        "Data documento",
        "Data pagamento",
        "Ivam",
        "Imponibile",
        "Totale fatture",
        "Identificativo SDI",
    ]
    OUTPUT_DATE_COLUMNS = ["Data emissione", "Data documento", "Data pagamento"]
    OUTPUT_MONEY_COLUMNS = ["Ivam", "Imponibile", "Totale fatture"]
    MAX_DETAIL_ROWS = 5000

    def process_file(self, input_path: Path, output_path: Path) -> Dict[str, Any]:
        try:
            logger.info("Caricamento file Pisa Ricevute: %s", input_path)
            try:
                df = pd.read_excel(input_path, usecols=self.INPUT_REQUIRED_COLUMNS, dtype=str)
            except ValueError:
                df_header = pd.read_excel(input_path, nrows=0)
                missing_columns = [col for col in self.INPUT_REQUIRED_COLUMNS if col not in df_header.columns]
                if missing_columns:
                    raise ValueError(f"Colonne mancanti: {', '.join(missing_columns)}")
                raise

            totale_fattura = pd.to_numeric(
                df["Importo fattura"].astype(str).str.replace(",", ".", regex=False),
                errors="coerce",
            ).fillna(0)
            iva = pd.to_numeric(
                df["IVA"].astype(str).str.replace(",", ".", regex=False),
                errors="coerce",
            ).fillna(0)
            imponibile = totale_fattura - iva

            df_finale = pd.DataFrame(
                {
                    "Ragione sociale": df["Creditore"],
                    "N.fatture": df["Numero fattura"],
                    "Data emissione": pd.to_datetime(df["Data emissione"], errors="coerce"),
                    "Data documento": pd.to_datetime(df["Data documento"], errors="coerce"),
                    "Data pagamento": pd.to_datetime(df["Data pagamento"], errors="coerce"),
                    "Ivam": iva,
                    "Imponibile": imponibile,
                    "Totale fatture": totale_fattura,
                    "Identificativo SDI": df["Identificativo SDI"],
                }
            )
            df_finale = df_finale[self.OUTPUT_COLUMNS]

            cartacee_df, elettroniche_df = self._split_by_sdi(df_finale, "Identificativo SDI")
            self._create_excel_output(df_finale, cartacee_df, elettroniche_df, output_path, display_df=df_finale)
            stats = {
                "total_records": len(df_finale),
                "fase2_records": len(cartacee_df),
                "fase3_records": len(elettroniche_df),
                "duplicates_removed": 0,
                "protocols_fase2": {"Cartacee": len(cartacee_df)},
                "protocols_fase3": {"Elettroniche": len(elettroniche_df)},
            }
            logger.info("File Pisa Ricevute elaborato con successo: %s", stats)
            return stats
        except Exception as exc:
            logger.error("Errore elaborazione file Pisa Ricevute: %s", str(exc))
            raise

    def _create_excel_output(
        self,
        df: pd.DataFrame,
        cartacee_df: pd.DataFrame,
        elettroniche_df: pd.DataFrame,
        output_path: Path,
        display_df: Optional[pd.DataFrame] = None,
    ) -> None:
        wb = Workbook()

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True)

        dati_df = display_df if display_df is not None else df
        self._add_dataframe_sheet(
            wb,
            "Dati",
            dati_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
            date_columns=self.OUTPUT_DATE_COLUMNS,
            date_format="dd/mm/yyyy",
            money_columns=self.OUTPUT_MONEY_COLUMNS,
            auto_size=False,
            use_active=True,
        )

        ws_cartacee = wb.create_sheet("Fatture Cartacee")
        self._create_simple_summary_sheet(
            ws_cartacee,
            cartacee_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        ws_elettroniche = wb.create_sheet("Fatture Elettroniche")
        self._create_simple_summary_sheet(
            ws_elettroniche,
            elettroniche_df,
            header_fill,
            header_font,
            total_fill,
            total_font,
        )

        wb.save(output_path)

    def _split_by_sdi(self, df: pd.DataFrame, sdi_column: str) -> tuple[pd.DataFrame, pd.DataFrame]:
        sdi_series = df[sdi_column]
        normalized = sdi_series.astype(str).str.strip()
        normalized_lower = normalized.str.lower()
        zero_mask = normalized_lower.str.fullmatch(r"0+(\.0+)?").fillna(False)
        empty_mask = normalized_lower.isin(["", "nan", "none", "null"]) | zero_mask
        cartacee_df = df[empty_mask].copy()
        elettroniche_df = df[~empty_mask].copy()
        return cartacee_df, elettroniche_df

    def _create_simple_summary_sheet(
        self,
        ws,
        df: pd.DataFrame,
        header_fill: PatternFill,
        header_font: Font,
        total_fill: PatternFill,
        total_font: Font,
    ) -> None:
        ws["A1"] = "NUMERO TOTALE"
        ws["B1"] = "TOTALE FATTURE"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        totale_fatture = pd.to_numeric(df["Totale fatture"], errors="coerce").sum()
        ws["A2"] = len(df)
        ws["B2"] = totale_fatture
        ws["B2"].number_format = "#,##0.00"

        for cell in ws[2]:
            cell.fill = total_fill
            cell.font = total_font

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20


class CompareFTFileProcessor(NFSFTFileProcessor):
    SELECTED_LETTERS = PisaFTFileProcessor.SELECTED_LETTERS
    RENAME_MAP = PisaFTFileProcessor.RENAME_MAP

    def process_files(self, nfs_path: Path, pisa_path: Path, output_path: Path) -> Dict[str, Any]:
        nfs_df = self._prepare_nfs_df(nfs_path)
        nfs_cartacee, nfs_elettroniche = self._split_by_sdi(nfs_df, "Identificativo SDI")

        pisa_df = self._prepare_pisa_df(pisa_path)
        pisa_jan = self._filter_january_2025(pisa_df, self._data_pagamento_column_name)
        pisa_cartacee, pisa_elettroniche = self._split_by_sdi(pisa_jan, self._sdi_column_name)
        pisa_importo_column = self._get_pisa_importo_column(pisa_jan)
        pisa_non_in_nfs = self._filter_pisa_non_in_nfs(pisa_elettroniche, nfs_elettroniche, self._sdi_column_name)
        pisa_diff = self._dedupe_by_sdi(pisa_non_in_nfs, self._sdi_column_name)

        nfs_elettroniche_count = self._count_unique_sdi(nfs_elettroniche, "Identificativo SDI")
        pisa_elettroniche_count = self._count_unique_sdi(pisa_elettroniche, self._sdi_column_name)

        summary = {
            "period": "2025-01",
            "nfs": {
                "cartacee": {
                    "count": len(nfs_cartacee),
                    "imponibile": pd.to_numeric(nfs_cartacee["Tot. Imponibile"], errors="coerce").sum(),
                },
                "elettroniche": {
                    "count": nfs_elettroniche_count,
                    "imponibile": pd.to_numeric(nfs_elettroniche["Tot. Imponibile"], errors="coerce").sum(),
                },
            },
            "pisa": {
                "cartacee": {
                    "count": len(pisa_cartacee),
                    "imponibile": pd.to_numeric(pisa_cartacee[pisa_importo_column], errors="coerce").sum(),
                },
                "elettroniche": {
                    "count": pisa_elettroniche_count,
                    "imponibile": pd.to_numeric(pisa_elettroniche[pisa_importo_column], errors="coerce").sum(),
                },
            },
        }

        self._create_compare_output(summary, pisa_diff, output_path)
        return summary

    def _prepare_nfs_df(self, input_path: Path) -> pd.DataFrame:
        df = pd.read_excel(input_path, dtype=str)
        self.validate_file(df)
        df["FAT_PROT"] = df["FAT_PROT"].astype(str).str.strip()
        df_senza_duplicati = df.drop_duplicates(subset=["FAT_NUM", "C_NOME"]).copy()
        df_filtrato = df_senza_duplicati[df_senza_duplicati["FAT_PROT"].isin(self.all_protocols)].copy()
        if len(df_filtrato) == 0:
            raise ValueError("Nessun protocollo valido trovato nel file NFS")
        df_filtrato["RA_CODTRIB"] = (
            df_filtrato["RA_CODTRIB"]
            .astype(str)
            .str.strip()
            .where(lambda value: value.isin(["I9", "RO"]), "")
        )
        colonne_ordinate = [
            "C_NOME",
            "FAT_DATDOC",
            "FAT_NDOC",
            "FAT_DATREG",
            "FAT_PROT",
            "FAT_NUM",
            "FAT_TOTIVA",
            "IMPONIBILE",
            "FAT_TOTFAT",
            "RA_CODTRIB",
            "RA_IMPOSTA",
            "RA_IMPON",
            "TMC_G8",
        ]
        df_finale = df_filtrato[colonne_ordinate].copy()
        df_finale.columns = [
            "Ragione Sociale",
            "Data Fatture",
            "N. Fatture",
            "Data Registrazione",
            "Protocollo",
            "N. Protocollo",
            "Imposta",
            "Tot. Imponibile",
            "Tot. Imp. Fatture",
            "Rit. Codice Tributo",
            "Rit. Imposta",
            "Rit. Imp.",
            "Identificativo SDI",
        ]
        df_finale["Data Registrazione"] = pd.to_datetime(df_finale["Data Registrazione"], errors="coerce")
        return df_finale

    def _prepare_pisa_df(self, input_path: Path) -> pd.DataFrame:
        df = pd.read_excel(input_path)
        required_indices = self._letters_to_indices(self.SELECTED_LETTERS)
        max_index = max(required_indices)
        if df.shape[1] <= max_index:
            missing_letters = [
                letter for letter, index in zip(self.SELECTED_LETTERS, required_indices) if index >= df.shape[1]
            ]
            raise ValueError(f"Colonne mancanti: {', '.join(missing_letters)}")

        data_pagamento_column = df.columns[self._letters_to_indices(["F"])[0]]
        pagamento_series = df[data_pagamento_column]
        pagamento_mask = ~(pagamento_series.isna() | (pagamento_series.astype(str).str.strip() == ""))
        df_pagato = df[pagamento_mask].copy()
        selected_indices = self._letters_to_indices(self.SELECTED_LETTERS)
        selected_columns = []
        for letter, index in zip(self.SELECTED_LETTERS, selected_indices):
            selected_columns.append(self.RENAME_MAP.get(letter) or df_pagato.columns[index])
        self._data_pagamento_column_name = self.RENAME_MAP.get("F") or data_pagamento_column
        self._sdi_column_name = selected_columns[self.SELECTED_LETTERS.index("A")]

        df_finale = df_pagato.iloc[:, selected_indices].copy()
        df_finale.columns = selected_columns

        return df_finale

    def _split_by_sdi(self, df: pd.DataFrame, sdi_column: str) -> tuple[pd.DataFrame, pd.DataFrame]:
        sdi_series = df[sdi_column]
        empty_mask = sdi_series.isna() | (sdi_series.astype(str).str.strip() == "")
        cartacee_df = df[empty_mask].copy()
        elettroniche_df = df[~empty_mask].copy()
        return cartacee_df, elettroniche_df

    def _letters_to_indices(self, letters: list[str]) -> list[int]:
        return [ord(letter) - ord("A") for letter in letters]

    def _get_pisa_importo_column(self, df: pd.DataFrame) -> str:
        for column in ("Importo Pagato", "Imponibile", "Imp.Tot. Fatture"):
            if column in df.columns:
                return column
        raise ValueError("Colonna importo non trovata nel file Pisa")

    def _filter_pisa_non_in_nfs(
        self,
        pisa_elettroniche: pd.DataFrame,
        nfs_elettroniche: pd.DataFrame,
        sdi_column: str,
    ) -> pd.DataFrame:
        if sdi_column not in pisa_elettroniche.columns or "Identificativo SDI" not in nfs_elettroniche.columns:
            return pisa_elettroniche.copy()
        pisa_sdi = self._normalize_sdi_series(pisa_elettroniche[sdi_column])
        nfs_sdi = self._normalize_sdi_series(nfs_elettroniche["Identificativo SDI"])
        nfs_set = set(nfs_sdi[nfs_sdi != ""])
        mask = ~pisa_sdi.isin(nfs_set)
        return pisa_elettroniche[mask].copy()

    def _normalize_sdi_series(self, series: pd.Series) -> pd.Series:
        normalized = series.astype(str).str.strip()
        normalized = normalized.where(~series.isna(), "")
        normalized = normalized.str.replace(r"\.0$", "", regex=True)
        normalized = normalized.str.replace(r"\D", "", regex=True)
        return normalized

    def _dedupe_by_sdi(self, df: pd.DataFrame, sdi_column: str) -> pd.DataFrame:
        if sdi_column not in df.columns:
            return df.copy()
        normalized = self._normalize_sdi_series(df[sdi_column])
        return df.loc[~normalized.duplicated()].copy()

    def _count_unique_sdi(self, df: pd.DataFrame, sdi_column: str) -> int:
        if sdi_column not in df.columns:
            return 0
        normalized = self._normalize_sdi_series(df[sdi_column])
        return int(normalized[normalized != ""].nunique())

    def _filter_january_2025(self, df: pd.DataFrame, date_column: str) -> pd.DataFrame:
        if date_column not in df.columns:
            return df.iloc[0:0].copy()
        date_series = pd.to_datetime(df[date_column], errors="coerce")
        start = pd.Timestamp(year=2025, month=1, day=1)
        end = pd.Timestamp(year=2025, month=1, day=31)
        mask = date_series.between(start, end)
        return df[mask].copy()

    def _create_compare_output(
        self,
        summary: Dict[str, Any],
        pisa_non_in_nfs: pd.DataFrame,
        output_path: Path,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Confronto Gennaio 2025"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True)

        ws.append(["Fonte", "Tipo Fatture", "Numero Totale", "Imponibile"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        rows = [
            ["FT NFS Pagato", "Cartacee", summary["nfs"]["cartacee"]["count"], summary["nfs"]["cartacee"]["imponibile"]],
            [
                "FT NFS Pagato",
                "Elettroniche",
                summary["nfs"]["elettroniche"]["count"],
                summary["nfs"]["elettroniche"]["imponibile"],
            ],
            ["FT Pisa Pagato", "Cartacee", summary["pisa"]["cartacee"]["count"], summary["pisa"]["cartacee"]["imponibile"]],
            [
                "FT Pisa Pagato",
                "Elettroniche",
                summary["pisa"]["elettroniche"]["count"],
                summary["pisa"]["elettroniche"]["imponibile"],
            ],
        ]

        for row in rows:
            ws.append(row)

        money_format = "#,##0.00"
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = money_format

        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = total_font

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

        if pisa_non_in_nfs is not None:
            date_columns = [column for column in pisa_non_in_nfs.columns if "data" in str(column).lower()]
            money_columns = [
                column
                for column in pisa_non_in_nfs.columns
                if column in ("Importo Pagato", "Imponibile", "Imp.Tot. Fatture")
            ]
            self._add_dataframe_sheet(
                wb,
                "Diffrenze da Verificare",
                pisa_non_in_nfs,
                header_fill,
                header_font,
                total_fill,
                total_font,
                date_columns=date_columns,
                money_columns=money_columns,
                date_format="dd/mm/yyyy",
                add_totals=False,
            )
        wb.save(output_path)
